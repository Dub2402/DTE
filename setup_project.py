from Source.Core.Enums import MediaPath, ExcelFilesPath

from dublib.Methods.Filesystem import MakeRootDirectories
from dublib.CLI.TextStyler.FastStyler import FastStyler
from dublib.Methods.System import Clear

from typing import Literal
from pathlib import Path
import itertools
import shutil
from enum import Enum

from openpyxl.utils import get_column_letter
from openpyxl import Workbook

Clear()

BUILD_ERRORS = []

class BuildStatus(str, Enum):
	CREATED = "CREATED"
	SKIPPED = "SKIPPED"
	FAILED = "FAILED"

class BuildStatusInformer:
	"""Информатор статуса работы автосборщика в консоль."""

	width = 99     
	indent = 5
	status_pad = 11

	@staticmethod
	def generate_instruction_for_user():
		"""Генерирует текст в консоль об окончании сборки проекта."""

		replaces = {
			"$media_name_files": "\n► ".join([Path(name_file.value).name for name_file in MediaPath]),
			"$excel_name_files": "\n► ".join([Path(name_file.value).name for name_file in ExcelFilesPath]),
			"$folder_media": FastStyler(MediaPath.start.folder).decorate.bold,
			"$folder_excel": FastStyler(ExcelFilesPath.buddy.folder).decorate.bold
		}

		template_text = (
			FastStyler("Для завершения настройки проекта необходимо: \n").decorate.bold,
			"1. Заполнить переменные в .env.\n",
			"2. Заполнить данные в таблицах Excel, в папке $folder_media:\n\n► $excel_name_files\n",
			"3. Добавить изображения в папку $folder_excel:\n\n► $media_name_files",
		)
		text = "\n".join(template_text)

		for replace_element in replaces:
			text = text.replace(replace_element, replaces[replace_element])

		return text

	@staticmethod
	def status_with_border(text: str, border_place: Literal[1, 2]):
		"""Выводит текст с границей в консоль.

		:param text: Текст.
		:type text: str
		:param border_place: Место отрисовки границы: 1 – над текстом, 2 – под текстом.
		:type border_place: Literal[1, 2]
		:raises ValueError: Неверное место отрисовки границы.
		"""

		if border_place not in (1, 2): raise ValueError(border_place)
		border = ("─" * (BuildStatusInformer.width + BuildStatusInformer.status_pad))

		match border_place:
			case 1: text = border + "\n" + text
			case 2: text = text + "\n" + border

		print(text)

	@staticmethod
	def continue_building_status(text: str, status: str | None = "", indent: bool = False):
		"""Выводит статусы работы автосборщиком над запущенным процессом.

		:param text: Текст.
		:type text: str
		:param status: Статус.
		:type status: str | None
		:param indent: Статус необходимости отступа.
		:type indent: bool
		"""

		if status: status = "[" + status + "]" 

		current_indent = BuildStatusInformer.indent if indent else 0
		current_width = BuildStatusInformer.width - current_indent

		print(" " * current_indent + text + "." * max(0, current_width - len(text)) + status)

class ExcelTemplater:
	"""Генератор шаблона Excel-таблиц."""

	@classmethod
	def headers(self) -> tuple[str, ...]:
		"""Названия столбцов таблицы Excel."""

		genders = ("МУЖ", "ЖЕН")
		types = ("Ежедневные", "Разовые и день в день")
		numbers = ("1", "2")

		return tuple(f"{gender} {types} {numbers}" for gender, types, numbers in itertools.product(genders, types, numbers))

	def create_excel_files(self, file_path: Path) -> bool:
		"""Автоматически создает excel-файл по указанному пути, если он отсутствуют в папке.

		:param file_path: Путь к таблице, которую нужно создать.
		:type file_path: Path
		:return: Статус успешности создания Excel-файла.
		:rtype: bool
		"""
		
		wb = Workbook()
		ws = wb.active
		ws.title = "Лист 1"
		ws.append(self.headers())

		for col_num, header_text in enumerate(self.headers(), start=1):
			col_letter = get_column_letter(col_num)
			header_len = len(header_text)
			ws.column_dimensions[col_letter].width = max(header_len + 4, 12)

		try: 
			wb.save(file_path)
			return True
		
		except Exception as E: 
			BuildStatusInformer.continue_building_status(f"🏷 Шаблон «{file_path.name} не создан.»", BuildStatus.FAILED, True)
			BUILD_ERRORS.append(f"[ШАБЛОНЫ EXCEL]: Ошибка при создании «{file_path.name}». Ошибка: {E}")
			return False

def setup_directories():
	"""Создает необходимые директории проекта."""

	status = BuildStatus.CREATED

	if Path(MediaPath.start.folder).exists() and Path(ExcelFilesPath.buddy.folder).exists(): status = BuildStatus.SKIPPED

	try: MakeRootDirectories([MediaPath.start.folder, ExcelFilesPath.buddy.folder])
	except Exception as e: 
		status = BuildStatus.FAILED
		BUILD_ERRORS.append(f"[ДИРЕКТОРИИ]: Не удалось создать папки проекта. Ошибка: {e}")
		
	BuildStatusInformer.continue_building_status("📁 1. Cоздание необходимых директорий", status.value)

def setup_excel_templates():
	"""Проверяет и создает Excel-шаблоны."""

	BuildStatusInformer.continue_building_status("📊 2. Создание Excel-шаблонов.")
	
	indent = True
	for table_path in ExcelFilesPath:
		file_path = Path(table_path.value)
		text = f"🏷 Шаблон «{file_path.name}»"
		
		if file_path.exists(): status = BuildStatus.SKIPPED

		else: 
			if ExcelTemplater().create_excel_files(file_path): status = BuildStatus.CREATED

		BuildStatusInformer.continue_building_status(text, status.value, indent)

def setup_env_file():
	"""Настраивает конфигурационный файл .env на основе шаблона .env.example."""
	
	env_path = Path(".env")
	example_path = Path(".env.example")
	
	if not env_path.exists():
		if example_path.exists():
			shutil.copy(example_path, env_path)
			status = BuildStatus.CREATED
		else:
			status = BuildStatus.FAILED
			BUILD_ERRORS.append("[КОНФИГУРАЦИЯ]: Не удалось создать «.env», так как в корне репозитория не найден шаблонный файл .env.example.")
	else:
		status = BuildStatus.SKIPPED
	
	example_path.unlink(missing_ok=True)
	BuildStatusInformer.continue_building_status("📝 3. Добавление конфигурации .env", status.value)

if __name__ == "__main__":

	BuildStatusInformer.status_with_border("👨‍💻 Начало автосборки проекта.", 2)

	setup_directories()
	setup_excel_templates()
	setup_env_file()

	if not BUILD_ERRORS: BuildStatusInformer.status_with_border("🚀 Сборка успешно завершена!", 1)
	else: 
		BuildStatusInformer.status_with_border(FastStyler("❌ Проблемы при сборке:\n").decorate.bold, 1)
		for error in BUILD_ERRORS:
			print(error)

	BuildStatusInformer.status_with_border(BuildStatusInformer.generate_instruction_for_user(), 1)