from pathlib import Path
import itertools

from openpyxl.utils import get_column_letter
from openpyxl import Workbook

class ExcelTemplater:
	"""Генератор шаблона Excel-таблиц."""

	@classmethod
	def headers(cls) -> tuple[str, ...]:
		"""Названия столбцов таблицы Excel."""

		genders = ("МУЖ", "ЖЕН")
		types = ("Ежедневные", "Разовые и день в день")
		numbers = ("1", "2")

		return tuple(f"{gender} {types} {numbers}" for gender, types, numbers in itertools.product(genders, types, numbers))

	@classmethod
	def create_excel_files(cls, file_path: Path) -> bool:
		"""Автоматически создает excel-файл по указанному пути, если он отсутствуют в папке.

		:param file_path: Путь к таблице, которую нужно создать.
		:type file_path: Path
		:return: Статус успешности создания Excel-файла.
		:rtype: bool
		"""
		
		wb = Workbook()
		ws = wb.active
		ws.title = "Лист 1"
		ws.append(cls.headers())

		for col_num, header_text in enumerate(cls.headers(), start=1):
			col_letter = get_column_letter(col_num)
			header_len = len(header_text)
			ws.column_dimensions[col_letter].width = max(header_len + 4, 12)
		wb.save(file_path)